import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

# Define os caminhos para entrada e saída
entrada_path = r"C:\Empresa\Dados\entrada"
objetivo_path = r"C:\Empresa\Dados\objetivo"
output_path = r"C:\Empresa\Resultados\Planilha_Atualizada.xlsx"  # Caminho completo do arquivo de saída

# Função para encontrar o arquivo Excel base na pasta de entrada
def encontrar_arquivo_base(pasta):
    for file in os.listdir(pasta):
        if file.endswith(".xlsx"):
            return os.path.join(pasta, file)
    raise FileNotFoundError(f"Nenhum arquivo Excel encontrado na pasta {pasta}.")

# Função para carregar os arquivos base de dados e objetivo, filtrando "Total Consolidado"
def carregar_dados():
    # Carregar o arquivo base de dados
    base_file = encontrar_arquivo_base(entrada_path)
    base_df = pd.read_excel(base_file, sheet_name="Dados")

    # Remover qualquer linha onde 'SETOR' contenha 'Total Consolidado'
    base_df = base_df[base_df['SETOR'] != 'Total Consolidado']
    
    # Carregar o arquivo de objetivo
    objetivo_file = encontrar_arquivo_base(objetivo_path)
    objetivo_df = pd.read_excel(objetivo_file, sheet_name="Metas")
    
    return base_df, objetivo_df

# Função para criar e preencher a planilha automaticamente
def criar_e_preencher_planilha():
    # Carregar os dados das planilhas
    base_df, objetivo_df = carregar_dados()
    
    # Criar a estrutura da aba "simulacao" e mesclar com o objetivo
    simulacao_df = pd.DataFrame({
        'SETOR': base_df['SETOR'],
        'EM ANDAMENTO': base_df['Em_Andamento'],
        'APROV. FINANCEIRA': base_df['Aprov_Financeira'],
        'APROV. GERAL': base_df['Aprov_Geral'],
        'CONCLUÍDO': base_df['Concluido']  # Usa 'Concluido' para preencher 'CONCLUÍDO'
    })
    
    # Realizar merge para obter 'META' usando 'SETOR' como chave
    if 'SETOR' in objetivo_df.columns and 'META' in objetivo_df.columns:
        simulacao_df = simulacao_df.merge(objetivo_df[['SETOR', 'META']], on='SETOR', how='left')
        simulacao_df['META'] = simulacao_df['META'].astype(float).fillna(0)
    else:
        simulacao_df['META'] = 0

    # Conversão de colunas para float e preenchimento de valores nulos
    simulacao_df['CONCLUÍDO'] = simulacao_df['CONCLUÍDO'].astype(float).fillna(0)
    simulacao_df['EM ANDAMENTO'] = simulacao_df['EM ANDAMENTO'].astype(float).fillna(0)
    simulacao_df['APROV. FINANCEIRA'] = simulacao_df['APROV. FINANCEIRA'].astype(float).fillna(0)
    simulacao_df['APROV. GERAL'] = simulacao_df['APROV. GERAL'].astype(float).fillna(0)

    # Calcular TOTAL como a soma de CONCLUÍDO, EM ANDAMENTO, APROV. FINANCEIRA e APROV. GERAL
    simulacao_df['TOTAL'] = (
        simulacao_df['CONCLUÍDO'] +
        simulacao_df['EM ANDAMENTO'] +
        simulacao_df['APROV. FINANCEIRA'] +
        simulacao_df['APROV. GERAL']
    )

    # Calcular DIFERENÇA META como a diferença entre META e TOTAL, com sinal negativo
    simulacao_df['DIFERENÇA META'] = simulacao_df.apply(
        lambda row: f"-{abs(row['META'] - row['TOTAL'])}", axis=1
    )

    # Calcular % ALCANCE como a divisão entre TOTAL e META e formatar como percentual
    simulacao_df['% ALCANCE'] = simulacao_df.apply(
        lambda row: f"{int((row['TOTAL'] / row['META']) * 100)}%" if row['META'] != 0 else None,
        axis=1
    )

    # Reordenar as colunas conforme solicitado
    simulacao_df = simulacao_df[
        ['SETOR', 'META', 'CONCLUÍDO', 'EM ANDAMENTO', 'APROV. FINANCEIRA', 
         'APROV. GERAL', 'TOTAL', 'DIFERENÇA META', '% ALCANCE']
    ]

    # Adicionar a linha Total Consolidado com somas e fórmulas específicas
    total_row = pd.DataFrame({
        'SETOR': ['Total Consolidado'],
        'META': [simulacao_df['META'].sum()],
        'CONCLUÍDO': [simulacao_df['CONCLUÍDO'].sum()],
        'EM ANDAMENTO': [simulacao_df['EM ANDAMENTO'].sum()],
        'APROV. FINANCEIRA': [simulacao_df['APROV. FINANCEIRA'].sum()],
        'APROV. GERAL': [simulacao_df['APROV. GERAL'].sum()],
        'TOTAL': [simulacao_df['TOTAL'].sum()],
        'DIFERENÇA META': [
            f"-{abs(simulacao_df['META'].sum() - simulacao_df['TOTAL'].sum())}"
        ],
        '% ALCANCE': [
            f"{int((simulacao_df['TOTAL'].sum() / simulacao_df['META'].sum()) * 100)}%"
            if simulacao_df['META'].sum() != 0 else None
        ]
    })

    # Concatenar a linha de total ao DataFrame final
    simulacao_df = pd.concat([simulacao_df, total_row], ignore_index=True)

    # Salvar o arquivo com estilo
    salvar_planilha_com_estilo(simulacao_df)

# Função para salvar a planilha com estilo visual e barras de dados
def salvar_planilha_com_estilo(simulacao_df):
    simulacao_df.to_excel(output_path, sheet_name='simulacao', index=False)

    wb = load_workbook(output_path)
    ws = wb['simulacao']

    # Definição de estilos
    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    unit_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    unit_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )
    alignment = Alignment(horizontal="center", vertical="center")

    # Formatar cabeçalhos
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Formatar coluna "SETOR"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1):
        for cell in row:
            cell.fill = unit_fill
            cell.font = unit_font
            cell.alignment = alignment
            cell.border = border

    # Formatar a linha "Total Consolidado"
    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.alignment = alignment
        cell.border = border

    # Aplicar bordas a todas as células
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # Adicionar barras de dados na coluna "% ALCANCE"
    percent_col_index = 9  # Índice da coluna "% ALCANCE"
    max_row = ws.max_row - 1  # Excluir a linha de total consolidado
    data_bar_rule = DataBarRule(
        start_type="num",
        start_value=0,
        end_type="num",
        end_value=100,
        color="63BE7B"  # Cor verde para as barras de dados
    )
    ws.conditional_formatting.add(f"I2:I{max_row}", data_bar_rule)

    wb.save(output_path)
    print(f"Planilha estilizada com barras de dados salva em: {output_path}")

# Executar o processo de criação e estilização da planilha
criar_e_preencher_planilha()